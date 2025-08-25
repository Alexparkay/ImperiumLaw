import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Load the existing spreadsheet
excel_file = 'output/PCBA_Import_Data.xlsx'
wb = load_workbook(excel_file)

# Define styles
header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
subheader_fill = PatternFill(start_color='95B3D7', end_color='95B3D7', fill_type='solid')
border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)

# Function to apply formatting to a worksheet
def format_worksheet(ws):
    # Format headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Format data cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = border
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

# Apply formatting to each worksheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    format_worksheet(ws)
    
    # Add table formatting if the sheet has data
    if ws.max_row > 1:
        tab = Table(displayName=f"Table_{sheet_name.replace(' ', '_')}", 
                   ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        
        # Add a default style
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        
        # Add the table to the worksheet
        try:
            ws.add_table(tab)
        except:
            # Table might already exist or other issue
            pass

# Add a title to the Overview sheet
ws = wb['Overview']
ws.insert_rows(1)
ws.merge_cells('A1:B1')
title_cell = ws['A1']
title_cell.value = "PCBA IMPORT DATA RESEARCH"
title_cell.font = Font(name='Arial', size=16, bold=True)
title_cell.alignment = Alignment(horizontal='center', vertical='center')

# Add filtering to relevant sheets
for sheet_name in ['HTS Codes', 'Top Importers', 'Top Suppliers', 'Import by Country', 'Industry Applications']:
    ws = wb[sheet_name]
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

# Save the formatted workbook
wb.save(excel_file)

print(f"PCBA Import Data spreadsheet formatting completed successfully at {os.path.abspath(excel_file)}")
