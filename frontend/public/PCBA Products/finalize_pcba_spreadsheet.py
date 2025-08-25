import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image

# Load the existing spreadsheet
excel_file = 'output/PCBA_Import_Data.xlsx'
wb = load_workbook(excel_file)

# Create a summary sheet
ws = wb.create_sheet("Summary", 0)

# Add title
ws['A1'] = "PCBA IMPORT DATA SUMMARY"
ws['A1'].font = Font(name='Arial', size=16, bold=True)
ws.merge_cells('A1:D1')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

# Add introduction
ws['A3'] = "This spreadsheet contains import data for Printed Circuit Board Assemblies (PCBAs) from China, Vietnam, Mexico, and Canada."
ws['A4'] = "The data is focused on companies with annual import volumes between $1 million and $10 million in the following target industries:"
ws['A5'] = "Medical, Oil & Gas, Metering, Green Energy, and Aerospace (excluding Automotive and Lighting sectors)."
ws['A6'] = "Use the tabs below to navigate through different aspects of the PCBA import data."

for row in range(3, 7):
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].alignment = Alignment(wrap_text=True)

# Add sheet navigation guide
ws['A8'] = "SHEET NAVIGATION GUIDE"
ws['A8'].font = Font(bold=True)
ws.merge_cells('A8:D8')
ws['A8'].alignment = Alignment(horizontal='center')

headers = ["Sheet Name", "Description", "Key Information"]
for col, header in enumerate(headers, 1):
    ws.cell(row=9, column=col).value = header
    ws.cell(row=9, column=col).font = Font(bold=True)
    ws.cell(row=9, column=col).alignment = Alignment(horizontal='center')

# Sheet descriptions
sheet_info = [
    ["Overview", "Research parameters and scope", "Product types, countries, industries, HTS codes"],
    ["HTS Codes", "Detailed HTS code information", "Code descriptions and relevance to PCBAs"],
    ["Top Importers", "Leading companies importing PCBAs", "Company names, locations, and potential industries"],
    ["Top Suppliers", "Leading companies supplying PCBAs", "Supplier names, countries, and notes"],
    ["Import by Country", "Import/export rankings by country", "Country rankings and target status"],
    ["Industry Applications", "PCBA applications by industry", "Specific applications and relevant HTS codes"],
    ["Data Limitations", "Limitations of publicly available data", "Challenges and recommendations"],
    ["Recommendations", "Suggested next steps", "Premium data sources and contact acquisition strategies"]
]

for row, info in enumerate(sheet_info, 10):
    for col, value in enumerate(info, 1):
        ws.cell(row=row, column=col).value = value
        ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)

# Adjust column widths
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 40

# Add instructions for filtering
ws['A20'] = "FILTERING INSTRUCTIONS"
ws['A20'].font = Font(bold=True)
ws.merge_cells('A20:D20')
ws['A20'].alignment = Alignment(horizontal='center')

filter_instructions = [
    ["1.", "Click on the filter button (funnel icon) in the column header", "Enables filtering for that column"],
    ["2.", "Use the dropdown menu to select specific values", "Shows only rows matching your criteria"],
    ["3.", "Multiple filters can be applied across different columns", "Narrows results based on combined criteria"],
    ["4.", "Clear filters by selecting 'Clear Filter' in the dropdown", "Returns to showing all data"]
]

for row, info in enumerate(filter_instructions, 21):
    for col, value in enumerate(info, 1):
        ws.cell(row=row, column=col).value = value
        ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)

# Add note about data sources
ws['A26'] = "NOTE: This data is compiled from publicly available sources including Descartes Datamyne, USITC Harmonized Tariff Schedule, International Trade Administration, and ImportGenius. For more detailed and company-specific information, premium database subscriptions are recommended as outlined in the 'Recommendations' sheet."
ws.merge_cells('A26:D28')
ws['A26'].alignment = Alignment(wrap_text=True)

# Apply borders to all cells with content
max_row = 28
max_col = 4
border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)

for row in range(1, max_row + 1):
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).border = border

# Add alternating row colors for readability
light_fill = PatternFill(start_color='EBF1DE', end_color='EBF1DE', fill_type='solid')
for row in range(10, 18, 2):  # Sheet info rows
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = light_fill

for row in range(21, 25, 2):  # Filter instruction rows
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = light_fill

# Save the updated workbook
wb.save(excel_file)

print(f"PCBA Import Data spreadsheet finalized successfully at {os.path.abspath(excel_file)}")
